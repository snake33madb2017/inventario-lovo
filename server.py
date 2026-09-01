import os
import sqlite3
from datetime import datetime
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import Workbook
import json
import time
import bcrypt
import smtplib
import ssl
from email.message import EmailMessage
import jwt
from datetime import timedelta
from fastapi import Depends, UploadFile, File, Form
from typing import Optional

app = FastAPI(title="Inventario Bar API")

SECRET_KEY = "super_secreto_lovo_2026_cambiar_en_prod"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 # 24 horas

import hashlib

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        if bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8')):
            return True
    except Exception:
        pass
    
    try:
        sha256_hash = hashlib.sha256(plain_password.encode('utf-8')).hexdigest()
        if sha256_hash == hashed_password:
            return True
    except Exception:
        pass
        
    return False

def get_password_hash(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import os
# Si existe la carpeta /data (Disco de Render), usamos esa ruta. Si no, ruta local.
DB_DIR = "/data" if os.path.exists("/data") else "."
DB_FILE = os.environ.get("DB_FILE", os.path.join(DB_DIR, "inventario.db"))

import shutil
if DB_DIR == "/data" and not os.path.exists(DB_FILE) and os.path.exists("inventario.db"):
    shutil.copy2("inventario.db", DB_FILE)

class Registro(BaseModel):
    categoria: str
    producto: str
    cantidad_dictada: float
    usuario: str = "Desconocido"

class LoginRequest(BaseModel):
    dni: str
    password: str

class NuevoUsuario(BaseModel):
    dni: str
    nombre: str
    password: str
    rol: str

class NuevaCategoria(BaseModel):
    nombre: str

class NuevoDiccionario(BaseModel):
    alias: str
    real_name: str

class NuevaReceta(BaseModel):
    nombre: str
    ingredientes: str
    procedimiento: str
    coste: str
    categoria: str

class AjusteBalance(BaseModel):
    producto: str
    stock_anterior: float
    stock_actual: float
    precio: float

last_registro_time = 0.0
last_registro_payload = ""

def get_db_connection():
    conn = sqlite3.connect(DB_FILE, timeout=10.0)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Registros
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            hora TEXT,
            categoria TEXT,
            producto TEXT,
            cantidad_dictada REAL,
            botellas_llenas INTEGER,
            restante_porcentaje TEXT,
            usuario TEXT
        )
    ''')
    # Usuarios
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dni TEXT UNIQUE,
            nombre TEXT,
            password TEXT,
            rol TEXT
        )
    ''')
    # Categorias
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE
        )
    ''')
    # Diccionario
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS diccionario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alias TEXT UNIQUE,
            real_name TEXT
        )
    ''')
    # Recetas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS recetas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE,
            ingredientes TEXT,
            procedimiento TEXT,
            coste TEXT,
            categoria TEXT
        )
    ''')
    
    # Stock de Referencia
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS stock_referencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto TEXT UNIQUE,
            categoria TEXT,
            stock_anterior REAL,
            precio_unitario REAL,
            stock_ideal REAL DEFAULT 0.0
        )
    ''')
    
    # Defaults
    cursor.execute('SELECT COUNT(*) FROM usuarios')
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO usuarios (dni, nombre, password, rol) VALUES (?, ?, ?, ?)", 
                       ("Z3738848L", "Marco Daza", get_password_hash("Lovo2026*"), "encargado"))
    else:
        # Asegurar que el usuario encargado Z3738848L siempre exista
        cursor.execute('SELECT COUNT(*) FROM usuarios WHERE dni = ?', ("Z3738848L",))
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO usuarios (dni, nombre, password, rol) VALUES (?, ?, ?, ?)", 
                           ("Z3738848L", "Marco Daza", get_password_hash("Lovo2026*"), "encargado"))
                       
    cursor.execute('SELECT COUNT(*) FROM categorias')
    if cursor.fetchone()[0] == 0:
        default_cats = ["Cristalería", "Licores", "Refrescos", "Cervezas", "Destilados", "Zumos"]
        for cat in default_cats:
            cursor.execute("INSERT INTO categorias (nombre) VALUES (?)", (cat,))
            
    cursor.execute('SELECT COUNT(*) FROM diccionario')
    if cursor.fetchone()[0] == 0:
        default_dict = [("brugau", "Brugal"), ("barcelo", "Barceló"), ("tonica", "Tónica"), ("cocacola", "Coca-Cola")]
        for alias, real_name in default_dict:
            cursor.execute("INSERT INTO diccionario (alias, real_name) VALUES (?, ?)", (alias, real_name))

    conn.commit()
    
    # Migrations
    try:
        cursor.execute("ALTER TABLE stock_referencia ADD COLUMN stock_ideal REAL DEFAULT 0.0")
        conn.commit()
    except Exception:
        pass # Column already exists
        
    load_stock_referencia(conn)
    inicializar_stock_julio(conn)
    sincronizar_categorias(conn)
    conn.close()

def sincronizar_categorias(conn):
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT categoria FROM registros WHERE categoria IS NOT NULL AND categoria != ''")
        categorias_usadas = cursor.fetchall()
        for row in categorias_usadas:
            cat = row[0].strip()
            cursor.execute("SELECT COUNT(*) FROM categorias WHERE nombre = ?", (cat,))
            if cursor.fetchone()[0] == 0:
                cursor.execute("INSERT INTO categorias (nombre) VALUES (?)", (cat,))
        conn.commit()
    except Exception as e:
        print("Error sincronizando categorias:", e)

def inicializar_stock_julio(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM registros WHERE fecha = '31/07/2026'")
    if cursor.fetchone()[0] > 0:
        return "Ya existe"
        
    import os
    if not os.path.exists("STOCK JULIO.xlsx"):
        return "No existe el archivo"
        
    try:
        import pandas as pd
        import math
        
        cursor.execute("SELECT alias, real_name FROM diccionario")
        dic_rows = cursor.fetchall()
        diccionario = {row['alias'].lower(): row['real_name'] for row in dic_rows}

        def normalizar_producto(nombre_raw):
            if not isinstance(nombre_raw, str): return ""
            clean = nombre_raw.strip().lower()
            return diccionario.get(clean, nombre_raw.strip())
            
        def clean_quantity(val):
            if pd.isna(val) or val is None or str(val).strip() == "": return 0.0
            val_str = str(val).strip().upper().replace("O", "0").replace(",", ".")
            try: return max(0.0, float(val_str))
            except: return 0.0
            
        def split_bottles_and_pct(cantidad):
            botellas_llenas = int(math.floor(cantidad))
            decimal_part = round(cantidad - botellas_llenas, 3)
            pct_val = int(round(decimal_part * 100))
            return botellas_llenas, f"{pct_val}%"
            
        dfs = pd.read_excel("STOCK JULIO.xlsx", sheet_name=None)
        
        records = []
        
        if 'Stock lovo' in dfs:
            df = dfs['Stock lovo']
            current_cat = "General"
            for _, row in df.iterrows():
                col_b = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                col_c = row.iloc[2] if len(row) > 2 else None
                col_d = row.iloc[3] if len(row) > 3 else None
                
                if not col_b or col_b.lower() == "producto": continue
                
                if col_b.isupper() and len(col_b) > 1 and pd.isna(col_c) and pd.isna(col_d):
                    current_cat = col_b.title()
                    continue
                    
                cleaned_prod = normalizar_producto(col_b)
                qty = clean_quantity(col_c)
                botellas, pct = split_bottles_and_pct(qty)
                
                records.append(("31/07/2026", "00:00:00", current_cat, cleaned_prod, qty, botellas, pct, "Cierre Julio"))
                
        if 'Cristaleria' in dfs:
            df = dfs['Cristaleria']
            item_cols = [c for c in df.columns if "VASOS" in str(c).upper() or "PRODUCTO" in str(c).upper()]
            total_cols = [c for c in df.columns if "TOTAL" in str(c).upper()]
            if item_cols and total_cols:
                for _, row in df.iterrows():
                    item = str(row[item_cols[0]]).strip() if pd.notna(row[item_cols[0]]) else ""
                    if not item or item.lower() == "nan": continue
                    qty = clean_quantity(row[total_cols[0]])
                    botellas, pct = split_bottles_and_pct(qty)
                    records.append(("31/07/2026", "00:00:00", "Cristalería", normalizar_producto(item), qty, botellas, pct, "Cierre Julio"))
                    
        if 'Stock producciones' in dfs:
            df = dfs['Stock producciones']
            for _, row in df.iterrows():
                prod_b = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                qty_b = clean_quantity(row.iloc[2]) if len(row) > 2 else 0.0
                if prod_b and prod_b.lower() not in ["nan", "producto en botella", "producciones"]:
                    botellas, pct = split_bottles_and_pct(qty_b)
                    records.append(("31/07/2026", "00:00:00", "Producción Botellas", normalizar_producto(prod_b), qty_b, botellas, pct, "Cierre Julio"))
                    
                prod_g = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ""
                qty_g = clean_quantity(row.iloc[5]) if len(row) > 5 else 0.0
                if prod_g and prod_g.lower() not in ["nan", "garrafas"]:
                    botellas, pct = split_bottles_and_pct(qty_g)
                    records.append(("31/07/2026", "00:00:00", "Producción Garrafas", normalizar_producto(prod_g), qty_g, botellas, pct, "Cierre Julio"))
                    
            cursor.executemany("INSERT INTO registros (fecha, hora, categoria, producto, cantidad_dictada, botellas_llenas, restante_porcentaje, usuario) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", records)
            conn.commit()
            return "Éxito: " + str(len(records)) + " registros insertados."
    except Exception as e:
        print("Error inicializando stock julio:", e)
        return "Error: " + str(e)

@app.get("/api/debug/julio")
def debug_julio():
    try:
        conn = get_db_connection()
        res = inicializar_stock_julio(conn)
        conn.close()
        return {"resultado": res}
    except Exception as e:
        return {"error": str(e)}

def load_stock_referencia(conn):
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM stock_referencia')
    if cursor.fetchone()[0] > 0:
        return
        
    import os
    if not os.path.exists("STOCK JULIO.xlsx"):
        return
        
    try:
        import pandas as pd
        import math
        
        cursor.execute("SELECT alias, real_name FROM diccionario")
        dic_rows = cursor.fetchall()
        diccionario = {row['alias'].lower(): row['real_name'] for row in dic_rows}

        def normalizar_producto(nombre_raw):
            if not isinstance(nombre_raw, str): return ""
            clean = nombre_raw.strip().lower()
            return diccionario.get(clean, nombre_raw.strip())
            
        dfs = pd.read_excel("STOCK JULIO.xlsx", sheet_name=None)
        
        stock_dict = {}
        
        if 'Productos-precio' in dfs:
            df_precios = dfs['Productos-precio']
            for _, row in df_precios.iterrows():
                if len(row) > 3:
                    prod = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                    precio = row.iloc[3]
                    if prod and prod.lower() != 'producto' and prod.lower() != 'nan':
                        precio_val = float(precio) if pd.notna(precio) and isinstance(precio, (int, float)) else 0.0
                        stock_dict[normalizar_producto(prod)] = {'precio': precio_val, 'stock': 0.0, 'categoria': 'General'}
                        
        def process_qty(qty):
            try: return float(qty) if pd.notna(qty) else 0.0
            except: return 0.0
            
        if 'Stock lovo' in dfs:
            df = dfs['Stock lovo']
            current_cat = "Spirits / Licores"
            for _, row in df.iterrows():
                col_b = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
                col_c = row.iloc[2] if len(row) > 2 else None
                col_d = row.iloc[3] if len(row) > 3 else None
                
                if not col_b or col_b.lower() == "producto" or col_b.lower() == "nan": continue
                if col_b.isupper() and len(col_b) > 1 and pd.isna(col_c) and pd.isna(col_d):
                    current_cat = col_b.title()
                    continue
                    
                prod = normalizar_producto(col_b)
                qty = process_qty(col_c)
                if prod not in stock_dict:
                    stock_dict[prod] = {'precio': 0.0, 'stock': 0.0, 'categoria': current_cat}
                stock_dict[prod]['stock'] += qty
                stock_dict[prod]['categoria'] = current_cat
                
        if 'Cristaleria' in dfs:
            df = dfs['Cristaleria']
            for _, row in df.iterrows():
                if len(row) > 1:
                    prod = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                    if not prod or prod.lower() == 'nan': continue
                    prod = normalizar_producto(prod)
                    total_cols = [c for c in df.columns if "TOTAL" in str(c).upper()]
                    qty = process_qty(row[total_cols[0]]) if total_cols else 0.0
                    precio_cols = [c for c in df.columns if "PRECIO" in str(c).upper()]
                    precio = process_qty(row[precio_cols[0]]) if precio_cols else 0.0
                    
                    if prod not in stock_dict:
                        stock_dict[prod] = {'precio': precio, 'stock': 0.0, 'categoria': 'Cristalería'}
                    elif stock_dict[prod]['precio'] == 0.0:
                        stock_dict[prod]['precio'] = precio
                    stock_dict[prod]['stock'] += qty
                    
        if 'Stock producciones' in dfs:
            df = dfs['Stock producciones']
            for _, row in df.iterrows():
                if len(row) > 2:
                    prod_b = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                    qty_b = process_qty(row.iloc[2])
                    if prod_b and prod_b.lower() not in ["nan", "producto en botella", "producciones"]:
                        prod_b = normalizar_producto(prod_b)
                        if prod_b not in stock_dict:
                            stock_dict[prod_b] = {'precio': 0.0, 'stock': 0.0, 'categoria': 'Producción Botellas'}
                        stock_dict[prod_b]['stock'] += qty_b
                if len(row) > 5:
                    prod_g = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ""
                    qty_g = process_qty(row.iloc[5])
                    if prod_g and prod_g.lower() not in ["nan", "garrafas"]:
                        prod_g = normalizar_producto(prod_g)
                        if prod_g not in stock_dict:
                            stock_dict[prod_g] = {'precio': 0.0, 'stock': 0.0, 'categoria': 'Producción Garrafas'}
                        stock_dict[prod_g]['stock'] += qty_g
                        
        insert_data = []
        for p, d in stock_dict.items():
            if d['stock'] > 0 or d['precio'] > 0:
                insert_data.append((p, d['categoria'], d['stock'], d['precio']))
            
        cursor.executemany("INSERT INTO stock_referencia (producto, categoria, stock_anterior, precio_unitario) VALUES (?, ?, ?, ?)", insert_data)
        conn.commit()
    except Exception as e:
        print(f"Error loading stock_referencia: {e}")

@app.on_event("startup")
def startup_event():
    init_db()

from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

security = HTTPBearer()

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        dni: str = payload.get("sub")
        if dni is None:
            raise HTTPException(status_code=401, detail="Token inválido")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token inválido")

def check_is_admin(user: dict = Depends(get_current_user)):
    if user.get("rol") != "encargado":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    return user

def check_is_produccion_or_admin(user: dict = Depends(get_current_user)):
    if user.get("rol") not in ["encargado", "produccion"]:
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    return user

@app.post("/api/login")
def login(req: LoginRequest):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        dni_clean = req.dni.strip().upper()
        cursor.execute('SELECT nombre, password, rol FROM usuarios WHERE dni = ?', (dni_clean,))
        user = cursor.fetchone()
        conn.close()
        
        if user and verify_password(req.password, user["password"]):
            access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
            access_token = create_access_token(
                data={"sub": dni_clean, "nombre": user["nombre"], "rol": user["rol"]}, expires_delta=access_token_expires
            )
            return {"status": "success", "nombre": user["nombre"], "rol": user["rol"], "token": access_token}
            
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ----- ENDPOINTS REGISTROS -----

@app.post("/api/registro")
def añadir_registro(registro: Registro, user: dict = Depends(get_current_user)):
    global last_registro_time, last_registro_payload
    payload_str = f"{registro.categoria}|{registro.producto}|{registro.cantidad_dictada}|{registro.usuario}"
    current_time = time.time()
    if payload_str == last_registro_payload and (current_time - last_registro_time) < 4.0:
        return {"status": "success", "message": "Registro duplicado ignorado por el servidor"}
    last_registro_payload = payload_str
    last_registro_time = current_time

    try:
        now = datetime.now()
        fecha = now.strftime("%d/%m/%Y")
        hora = now.strftime("%H:%M:%S")

        botellas_llenas = int(registro.cantidad_dictada)
        restante_porcentaje = round((registro.cantidad_dictada - botellas_llenas) * 100)

        if restante_porcentaje > 0 or "." in str(registro.cantidad_dictada):
            restante_str = f"{restante_porcentaje}%"
        else:
            botellas_llenas = int(registro.cantidad_dictada)
            restante_str = "-"
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Safety net: si viene sin categoría, intentar buscarla o poner General
        cat_final = registro.categoria
        if not cat_final or cat_final.strip() == "":
            cursor.execute('SELECT categoria FROM stock_referencia WHERE producto = ?', (registro.producto,))
            row = cursor.fetchone()
            if row:
                cat_final = row['categoria']
            else:
                cat_final = "General"

        cursor.execute('''
            INSERT INTO registros (fecha, hora, categoria, producto, cantidad_dictada, botellas_llenas, restante_porcentaje, usuario)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (fecha, hora, cat_final, registro.producto, registro.cantidad_dictada, botellas_llenas, restante_str, registro.usuario))
        conn.commit()
        conn.close()
        return {"status": "success", "message": "Registro añadido correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/registro/ultimo")
def borrar_ultimo(user: dict = Depends(get_current_user)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT id FROM registros WHERE usuario = ? ORDER BY id DESC LIMIT 1', (user.get("nombre"),))
            
        row = cursor.fetchone()
        if row:
            cursor.execute('DELETE FROM registros WHERE id = ?', (row['id'],))
            conn.commit()
            conn.close()
            return {"status": "success", "message": "Último registro eliminado"}
        conn.close()
        return {"status": "warning", "message": "No hay registros tuyos para borrar"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/inventario/todo")
def borrar_todo_inventario(user: dict = Depends(check_is_admin)):
    try:
        now = datetime.now()
        fecha_hoy = now.strftime("%d/%m/%Y")
        mes_año_actual = now.strftime("%m/%Y")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Obtener todos los registros de hoy para calcular el stock final
        cursor.execute('SELECT * FROM registros WHERE fecha = ?', (fecha_hoy,))
        rows = cursor.fetchall()
        
        stock_act = {}
        cat_map = {}
        for row in rows:
            prod = row['producto']
            cat = row['categoria']
            cat_map[prod] = cat
            b = row['botellas_llenas']
            r_str = row['restante_porcentaje']
            r_val = 0.0
            if r_str and r_str != '-':
                try: r_val = float(str(r_str).replace('%','')) / 100.0
                except: pass
            
            if prod not in stock_act: stock_act[prod] = 0.0
            stock_act[prod] += (b + r_val)
            
        # 2. Actualizar el stock_referencia para que el mes siguiente se compare con el cierre de hoy
        cursor.execute('UPDATE stock_referencia SET stock_anterior = 0.0')
        
        for prod, qty in stock_act.items():
            cat = cat_map[prod]
            cursor.execute('SELECT id FROM stock_referencia WHERE producto = ?', (prod,))
            if cursor.fetchone():
                cursor.execute('UPDATE stock_referencia SET stock_anterior = ? WHERE producto = ?', (qty, prod))
            else:
                cursor.execute('INSERT INTO stock_referencia (producto, categoria, stock_anterior, precio_unitario) VALUES (?, ?, ?, 0.0)', (prod, cat, qty))
        
        # 3. Borrar conteos de días anteriores de ESTE MES, manteniendo el de hoy (Cierre) y los meses pasados en el historial
        cursor.execute('DELETE FROM registros WHERE fecha LIKE ? AND fecha != ?', (f'%/{mes_año_actual}', fecha_hoy))
        
        conn.commit()
        conn.close()
        return {"status": "success", "message": f"Cierre guardado para el {fecha_hoy} y stock de referencia actualizado."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/balance/ajuste")
def ajustar_balance(ajuste: AjusteBalance, user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Update stock_referencia (Stock Anterior & Precio)
        cursor.execute('SELECT id FROM stock_referencia WHERE producto = ?', (ajuste.producto,))
        if cursor.fetchone():
            cursor.execute('UPDATE stock_referencia SET stock_anterior = ?, precio_unitario = ? WHERE producto = ?', 
                          (ajuste.stock_anterior, ajuste.precio, ajuste.producto))
        else:
            # If not in reference, add it
            # Category might be unknown if not in stock_referencia
            cat = 'Sin Categoría'
            cursor.execute('INSERT INTO stock_referencia (producto, categoria, stock_anterior, precio_unitario) VALUES (?, ?, ?, ?)',
                          (ajuste.producto, cat, ajuste.stock_anterior, ajuste.precio))
                          
        # 3. Update registros for Stock Actual
        # Delete today's entries for this product
        now = datetime.now()
        fecha_hoy = now.strftime("%d/%m/%Y")
        hora = now.strftime("%H:%M:%S")
        
        cursor.execute('DELETE FROM registros WHERE fecha = ? AND producto = ?', (fecha_hoy, ajuste.producto))
        
        # Insert a new clean entry with the exact adjusted stock
        if ajuste.stock_actual > 0 or True:
            # We insert it even if 0, so the adjustment is recorded
            cat_row = cursor.execute('SELECT categoria FROM stock_referencia WHERE producto = ?', (ajuste.producto,)).fetchone()
            cat = cat_row['categoria'] if cat_row else 'Sin Categoría'
            
            # Since botellas_llenas is int and restante is string, we'll store everything in botellas_llenas as float for the adjustment
            # Or split it:
            enteras = int(ajuste.stock_actual)
            decimal = int(round((ajuste.stock_actual - enteras) * 100))
            decimal_str = f"{decimal}%" if decimal > 0 else "-"
            
            cursor.execute('''
                INSERT INTO registros (fecha, hora, categoria, producto, cantidad_dictada, botellas_llenas, restante_porcentaje, usuario)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (fecha_hoy, hora, cat, ajuste.producto, ajuste.stock_actual, enteras, decimal_str, "Ajuste Manual"))
            
        conn.commit()
        conn.close()
        
        return {"status": "success", "message": "Ajuste manual aplicado correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inventario/hoy")
def obtener_inventario_hoy(user: dict = Depends(get_current_user)):
    try:
        now = datetime.now()
        fecha_hoy = now.strftime("%d/%m/%Y")
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM registros WHERE fecha = ? ORDER BY hora DESC LIMIT 150', (fecha_hoy,))
        rows = cursor.fetchall()
        conn.close()
        
        registros = []
        for row in rows:
            registros.append(dict(row))
        return {"registros": registros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class AjusteIdeal(BaseModel):
    producto: str
    stock_ideal: float

@app.get("/api/catalogo")
def obtener_catalogo(user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT producto, categoria, stock_ideal FROM stock_referencia ORDER BY categoria, producto')
        rows = cursor.fetchall()
        conn.close()
        return [{"producto": r["producto"], "categoria": r["categoria"], "stock_ideal": r["stock_ideal"]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/catalogo/ideal")
def actualizar_stock_ideal(ajuste: AjusteIdeal, user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('UPDATE stock_referencia SET stock_ideal = ? WHERE producto = ?', (ajuste.stock_ideal, ajuste.producto))
        conn.commit()
        conn.close()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/productos")
def obtener_productos_historicos(user: dict = Depends(get_current_user)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT producto FROM registros')
        rows = cursor.fetchall()
        conn.close()
        lista = [r["producto"] for r in rows if r["producto"]]
        lista.sort()
        return {"productos": lista}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inventario/fechas")
def obtener_fechas_historial(user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT fecha FROM registros ORDER BY substr(fecha, 7, 4) || substr(fecha, 4, 2) || substr(fecha, 1, 2) DESC')
        rows = cursor.fetchall()
        conn.close()
        fechas = [r["fecha"] for r in rows if r["fecha"]]
        fechas_unicas = list(dict.fromkeys(fechas))
        return {"fechas": fechas_unicas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inventario/historial")
def obtener_historial_fecha(fecha: str, user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM registros WHERE fecha = ? ORDER BY categoria ASC, producto ASC', (fecha,))
        rows = cursor.fetchall()
        conn.close()
        registros = [dict(row) for row in rows]
        return {"registros": registros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/inventario/historial")
def borrar_fecha_historial(fecha: str, user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM registros WHERE fecha = ?', (fecha,))
        deleted = cursor.rowcount
        conn.commit()
        conn.close()
        if deleted == 0:
            raise HTTPException(status_code=404, detail=f"No se encontraron registros para la fecha {fecha}")
        return {"status": "success", "message": f"Se eliminaron {deleted} registros de la fecha {fecha}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inventario/referencia")
def obtener_referencia(user: dict = Depends(get_current_user)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM stock_referencia')
        rows = cursor.fetchall()
        conn.close()
        referencia = [dict(row) for row in rows]
        return {"referencia": referencia}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/inventario/comparativa")
def obtener_comparativa(fecha: Optional[str] = None, user: dict = Depends(check_is_admin)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if not fecha:
            cursor.execute('SELECT fecha FROM registros ORDER BY id DESC LIMIT 1')
            row = cursor.fetchone()
            if row:
                fecha = row['fecha']
            else:
                now = datetime.now()
                fecha = now.strftime("%d/%m/%Y")
        
        cursor.execute('SELECT * FROM stock_referencia')
        ref_rows = cursor.fetchall()
        stock_ref = {row['producto']: dict(row) for row in ref_rows}
        
        cursor.execute('SELECT * FROM registros WHERE fecha = ?', (fecha,))
        reg_rows = cursor.fetchall()
        
        stock_act = {}
        for row in reg_rows:
            prod = row['producto']
            if prod not in stock_act:
                stock_act[prod] = 0.0
            
            botellas = row['botellas_llenas']
            rest_str = row['restante_porcentaje']
            rest_val = 0.0
            if rest_str and rest_str != '-':
                try:
                    rest_str_clean = str(rest_str).strip()
                    if '%' in rest_str_clean:
                        rest_val = float(rest_str_clean.replace('%', '')) / 100.0
                    else:
                        rest_val = float(rest_str_clean)
                        # if someone entered 50 instead of 50%, we can assume it's percentage if > 1
                        if rest_val > 1:
                            rest_val = rest_val / 100.0
                except: pass
            
            stock_act[prod] += (botellas + rest_val)
            
        conn.close()
        
        comparativa = []
        todos = set(stock_ref.keys()) | set(stock_act.keys())
        
        for p in sorted(list(todos)):
            s_ant = stock_ref[p]['stock_anterior'] if p in stock_ref else 0.0
            s_act = stock_act[p] if p in stock_act else 0.0
            precio = stock_ref[p]['precio_unitario'] if p in stock_ref else 0.0
            cat = stock_ref[p]['categoria'] if p in stock_ref else 'Sin Categoría'
            
            consumo = s_ant - s_act
            coste = consumo * precio if consumo > 0 else 0.0
            
            alerta = "OK"
            if p not in stock_act and s_ant > 0:
                alerta = "No Contado"
            elif consumo < 0:
                alerta = "Stock Negativo"
            elif s_ant > 0 and consumo > (s_ant * 0.5):
                alerta = "Consumo Elevado"
                
            comparativa.append({
                "producto": p,
                "categoria": cat,
                "stock_anterior": round(s_ant, 2),
                "stock_actual": round(s_act, 2),
                "consumo": round(consumo, 2),
                "coste_consumo": round(coste, 2),
                "precio_unitario": precio,
                "alerta": alerta
            })
            
        return {"comparativa": comparativa}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/descargar/hoy")
def descargar_excel_hoy(fecha: Optional[str] = None, user: dict = Depends(check_is_admin)):
    try:
        from openpyxl import load_workbook
        import os
        from fastapi.responses import FileResponse
        
        if not fecha:
            now = datetime.now()
            fecha_busqueda = now.strftime("%d/%m/%Y")
            fecha_archivo = now.strftime("%Y-%m-%d")
        else:
            fecha_busqueda = fecha
            fecha_archivo = fecha.replace("/", "-")
            
        template_file = "STOCK JULIO.xlsx"
        if not os.path.exists(template_file):
            raise HTTPException(status_code=404, detail="Plantilla STOCK JULIO.xlsx no encontrada")
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Cargar diccionario
        cursor.execute("SELECT alias, real_name FROM diccionario")
        dic_rows = cursor.fetchall()
        diccionario = {r['alias'].lower(): r['real_name'].lower() for r in dic_rows}
        for r in dic_rows:
            diccionario[r['real_name'].lower()] = r['real_name'].lower()
            
        # Cargar todos los productos de stock referencia
        cursor.execute("SELECT producto FROM stock_referencia")
        for r in cursor.fetchall():
            prod_clean = r['producto'].strip().lower()
            if prod_clean not in diccionario:
                diccionario[prod_clean] = prod_clean

        # Cargar registros del día
        cursor.execute('SELECT * FROM registros WHERE fecha = ?', (fecha_busqueda,))
        rows = cursor.fetchall()
        conn.close()
        
        stock_act = {}
        auditors = {}
        for row in rows:
            prod_raw = row['producto'].strip().lower()
            prod_norm = diccionario.get(prod_raw, prod_raw)
            
            b = row['botellas_llenas']
            r_str = row['restante_porcentaje']
            r_val = 0.0
            if r_str and r_str != '-':
                try:
                    rest_str_clean = str(r_str).strip()
                    if '%' in rest_str_clean:
                        r_val = float(rest_str_clean.replace('%', '')) / 100.0
                    else:
                        r_val = float(rest_str_clean)
                        if r_val > 1:
                            r_val = r_val / 100.0
                except: pass
            total_qty = b + r_val
            if prod_norm not in stock_act: 
                stock_act[prod_norm] = 0.0
                auditors[prod_norm] = set()
            stock_act[prod_norm] += total_qty
            if row['usuario']:
                auditors[prod_norm].add(row['usuario'])
            
        # Modificar Excel
        wb = load_workbook(template_file)
        ignore_words = {"producto", "total", "precio", "articulos", "cristaleria", "producciones", "botellas", "garrafas", "observaciones"}
        
        for ws in wb.worksheets:
            auditor_col = ws.max_column + 1
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_norm = cell.value.strip().lower()
                        if cell_norm == "producto" or cell_norm == "articulos":
                            # Poner encabezado de auditores en la misma fila que el encabezado Producto
                            header_cell = ws.cell(row=cell.row, column=auditor_col)
                            header_cell.value = "Auditores"
                            from openpyxl.styles import Font
                            header_cell.font = Font(bold=True)
                        elif cell_norm not in ignore_words and cell_norm in diccionario:
                            real_prod = diccionario[cell_norm]
                            right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            # Si la celda derecha está vacía o ya es numérica, sobrescribimos
                            # Ya que "justo a la derecha" es donde se anotan las cantidades
                            # IMPORTANT: No sobrescribir fórmulas
                            if not (isinstance(right_cell.value, str) and right_cell.value.startswith('=')):
                                right_cell.value = stock_act.get(real_prod, 0.0)
                            
                            # Escribir auditor en la nueva columna al final de la tabla
                            if real_prod in auditors and auditors[real_prod]:
                                ws.cell(row=cell.row, column=auditor_col).value = ", ".join(auditors[real_prod])

        temp_file = f"Inventario_Cierre_{fecha_archivo}.xlsx"
        wb.save(temp_file)
        return FileResponse(path=temp_file, filename=temp_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/inventario/importar-excel")
async def importar_excel_inventario(fecha: str = Form(...), file: UploadFile = File(...), user: dict = Depends(check_is_admin)):
    try:
        contents = await file.read()
        import io
        import math
        import pandas as pd
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT alias, real_name FROM diccionario")
        dic_rows = cursor.fetchall()
        diccionario = {row['alias'].lower(): row['real_name'] for row in dic_rows}

        def normalizar_producto(nombre_raw):
            if not isinstance(nombre_raw, str): return ""
            clean = nombre_raw.strip().lower()
            return diccionario.get(clean, nombre_raw.strip())
            
        dfs = pd.read_excel(io.BytesIO(contents), sheet_name=None)
        
        nuevos_registros = []
        hojas_detectadas = []
        
        if 'Stock lovo' in dfs:
            hojas_detectadas.append('Stock lovo')
            df = dfs['Stock lovo']
            df.columns = df.columns.astype(str).str.strip().str.lower()
            if 'articulos' in df.columns and 'botellas llenas' in df.columns:
                for _, row in df.iterrows():
                    prod = row['articulos']
                    if pd.isna(prod): continue
                    producto_real = normalizar_producto(str(prod))
                    llenas = row['botellas llenas']
                    try:
                        llenas_val = int(llenas) if pd.notna(llenas) else 0
                    except:
                        llenas_val = 0
                    restante = row.get('botella empezada (%)', row.get('botella empezada'))
                    try:
                        restante_val = float(restante) * 100 if isinstance(restante, float) and pd.notna(restante) else (float(restante) if pd.notna(restante) else 0.0)
                        if math.isnan(restante_val): restante_val = 0.0
                    except:
                        restante_val = 0.0
                    
                    nuevos_registros.append((
                        fecha, "00:00", "Spirits / Licores", producto_real, 
                        f"Llenas: {llenas_val}, Restante: {restante_val}%", 
                        llenas_val, float(restante_val), user.get("nombre", "Sistema")
                    ))
            elif 'producto' in df.columns and 'total' in df.columns:
                for _, row in df.iterrows():
                    prod = row['producto']
                    if pd.isna(prod): continue
                    if pd.isna(row['total']): continue
                    producto_real = normalizar_producto(str(prod))
                    total_val = float(row['total']) if pd.notna(row['total']) else 0.0
                    llenas_val = int(total_val)
                    restante_val = (total_val - llenas_val) * 100
                    nuevos_registros.append((
                        fecha, "00:00", "Spirits / Licores", producto_real, 
                        f"Llenas: {llenas_val}, Restante: {restante_val:.0f}%", 
                        llenas_val, float(restante_val), user.get("nombre", "Sistema")
                    ))
        
        if 'Cristaleria' in dfs:
            hojas_detectadas.append('Cristaleria')
            df = dfs['Cristaleria']
            df.columns = df.columns.astype(str).str.strip().str.lower()
            col_prod = 'articulos' if 'articulos' in df.columns else ('cristaleria' if 'cristaleria' in df.columns else None)
            
            if col_prod and 'unidades' in df.columns:
                for _, row in df.iterrows():
                    prod = row[col_prod]
                    if pd.isna(prod): continue
                    producto_real = normalizar_producto(str(prod))
                    unidades = row['unidades']
                    try:
                        llenas_val = int(unidades) if pd.notna(unidades) else 0
                    except:
                        llenas_val = 0
                    
                    nuevos_registros.append((
                        fecha, "00:00", "Cristaleria / Vinos", producto_real, 
                        f"Unidades: {llenas_val}", 
                        llenas_val, 0.0, user.get("nombre", "Sistema")
                    ))
            elif 'producto' in df.columns and 'total' in df.columns:
                for _, row in df.iterrows():
                    prod = row['producto']
                    if pd.isna(prod): continue
                    if pd.isna(row['total']): continue
                    producto_real = normalizar_producto(str(prod))
                    llenas_val = int(float(row['total']) if pd.notna(row['total']) else 0.0)
                    nuevos_registros.append((
                        fecha, "00:00", "Cristaleria / Vinos", producto_real, 
                        f"Unidades: {llenas_val}", 
                        llenas_val, 0.0, user.get("nombre", "Sistema")
                    ))
                    
        if 'Stock producciones' in dfs:
            hojas_detectadas.append('Stock producciones')
            df = dfs['Stock producciones']
            df.columns = df.columns.astype(str).str.strip().str.lower()
            col_prod = 'articulos' if 'articulos' in df.columns else ('producciones' if 'producciones' in df.columns else None)
            
            if col_prod and 'unidades' in df.columns:
                for _, row in df.iterrows():
                    prod = row[col_prod]
                    if pd.isna(prod): continue
                    producto_real = normalizar_producto(str(prod))
                    unidades = row['unidades']
                    try:
                        llenas_val = int(unidades) if pd.notna(unidades) else 0
                    except:
                        llenas_val = 0
                    
                    nuevos_registros.append((
                        fecha, "00:00", "Producciones / Batch", producto_real, 
                        f"Unidades: {llenas_val}", 
                        llenas_val, 0.0, user.get("nombre", "Sistema")
                    ))
            elif 'producto' in df.columns and 'total' in df.columns:
                for _, row in df.iterrows():
                    prod = row['producto']
                    if pd.isna(prod): continue
                    if pd.isna(row['total']): continue
                    producto_real = normalizar_producto(str(prod))
                    llenas_val = int(float(row['total']) if pd.notna(row['total']) else 0.0)
                    nuevos_registros.append((
                        fecha, "00:00", "Producciones / Batch", producto_real, 
                        f"Unidades: {llenas_val}", 
                        llenas_val, 0.0, user.get("nombre", "Sistema")
                    ))
                    
        if nuevos_registros:
            cursor.executemany('''
                INSERT INTO registros (fecha, hora, categoria, producto, cantidad_dictada, botellas_llenas, restante_porcentaje, usuario)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', nuevos_registros)
            conn.commit()
            conn.close()
            return {"message": f"Excel importado. Hojas procesadas: {', '.join(hojas_detectadas)}. {len(nuevos_registros)} registros guardados."}
        else:
            conn.close()
            raise HTTPException(status_code=400, detail="No se encontraron datos compatibles en el Excel.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error importando Excel: {str(e)}")

@app.post("/api/admin/enviar-excel")
def enviar_excel_correo(user: dict = Depends(check_is_admin)):
    sender_email = os.getenv("SMTP_EMAIL")
    sender_password = os.getenv("SMTP_PASSWORD")
    receiver_email = os.getenv("SMTP_DESTINATION")

    if not sender_email or not sender_password or not receiver_email:
        raise HTTPException(status_code=400, detail="El correo no está configurado en el servidor (faltan variables SMTP).")

    try:
        now = datetime.now()
        fecha_hoy = now.strftime("%d/%m/%Y")
        fecha_archivo = now.strftime("%Y-%m-%d")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM registros WHERE fecha = ? ORDER BY categoria ASC, hora ASC', (fecha_hoy,))
        rows = cursor.fetchall()
        conn.close()
        
        wb = Workbook()
        wb.remove(wb.active)
        
        categorias_dict = {}
        for row in rows:
            cat = row["categoria"]
            if cat not in categorias_dict: categorias_dict[cat] = []
            categorias_dict[cat].append([row["fecha"], row["hora"], row["categoria"], row["producto"], row["cantidad_dictada"], row["botellas_llenas"], row["restante_porcentaje"], row["usuario"]])
            
        encabezados = ["Fecha", "Hora", "Categoría", "Producto", "Cantidad Dictada", "Botellas Llenas", "Restante (%)", "Usuario"]
        
        import re
        for cat, filas in categorias_dict.items():
            cat_safe = re.sub(r'[\\*?:/\[\]]', '', cat)[:31]
            if not cat_safe: cat_safe = "Categoria"
            ws = wb.create_sheet(title=cat_safe)
            ws.append(encabezados)
            for f in filas: ws.append(f)
                
        if len(wb.sheetnames) == 0:
             ws = wb.create_sheet(title="Vacio")
             ws.append(["No hay registros hoy"])
                
        temp_file = f"Inventario_Hoy_{fecha_archivo}.xlsx"
        wb.save(temp_file)

        # Enviar por correo
        msg = EmailMessage()
        msg['Subject'] = f'Inventario Lovo - {fecha_hoy}'
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg.set_content(f"Hola,\n\nAdjunto el inventario del día {fecha_hoy} generado automáticamente por el sistema.\n\nSaludos,\nInventario Coctelería Lovo")

        with open(temp_file, 'rb') as f:
            file_data = f.read()
            
        msg.add_attachment(file_data, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=temp_file)

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg)

        os.remove(temp_file)
        return {"status": "success", "message": "Correo enviado correctamente."}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")

# ----- ENDPOINTS ADMIN -----

@app.get("/api/admin/usuarios")
def get_usuarios(user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, dni, nombre, rol FROM usuarios')
    users = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return users

@app.post("/api/admin/usuarios")
def crear_usuario(u: NuevoUsuario, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO usuarios (dni, nombre, password, rol) VALUES (?, ?, ?, ?)", (u.dni.strip().upper(), u.nombre, get_password_hash(u.password), u.rol))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="El DNI ya existe")
    conn.close()
    return {"status": "success"}

@app.delete("/api/admin/usuarios/{uid}")
def borrar_usuario(uid: int, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM usuarios WHERE id = ?', (uid,))
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.get("/api/admin/categorias")
def get_categorias(user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nombre FROM categorias')
    cats = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return cats

@app.post("/api/admin/categorias")
def crear_categoria(c: NuevaCategoria, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO categorias (nombre) VALUES (?)", (c.nombre,))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="La categoría ya existe")
    conn.close()
    return {"status": "success"}

@app.delete("/api/admin/categorias/{cid}")
def borrar_categoria(cid: int, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM categorias WHERE id = ?', (cid,))
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.get("/api/admin/diccionario")
def get_diccionario(user: dict = Depends(get_current_user)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, alias, real_name FROM diccionario')
        rows = cursor.fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/diccionario")
def crear_diccionario(d: NuevoDiccionario, user: dict = Depends(get_current_user)):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO diccionario (alias, real_name) VALUES (?, ?)", (d.alias.lower(), d.real_name))
        conn.commit()
        conn.close()
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/admin/diccionario/{did}")
def borrar_diccionario(did: int, user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM diccionario WHERE id = ?', (did,))
    conn.commit()
    conn.close()
    return {"status": "success"}

# ----- ENDPOINTS LABORATORIO / RECETAS -----

@app.get("/api/recetas")
def get_recetas(user: dict = Depends(check_is_produccion_or_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM recetas')
    rows = cursor.fetchall()
    conn.close()
    recetas = []
    for r in rows:
        receta = dict(r)
        if user.get("rol") != "encargado":
            receta["coste"] = "***" # Ocultar coste
        recetas.append(receta)
    return recetas

@app.post("/api/recetas")
def crear_receta(r: NuevaReceta, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO recetas (nombre, ingredientes, procedimiento, coste, categoria) VALUES (?, ?, ?, ?, ?)", 
                       (r.nombre, r.ingredientes, r.procedimiento, r.coste, r.categoria))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="La receta ya existe")
    conn.close()
    return {"status": "success"}

@app.delete("/api/recetas/{rid}")
def borrar_receta(rid: int, user: dict = Depends(check_is_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM recetas WHERE id = ?', (rid,))
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.post("/api/recetas/{rid}/producir")
def producir_receta(rid: int, user: dict = Depends(check_is_produccion_or_admin)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT nombre, categoria FROM recetas WHERE id = ?', (rid,))
    receta = cursor.fetchone()
    if not receta:
        conn.close()
        raise HTTPException(status_code=404, detail="Receta no encontrada")
    
    # Registrar en el inventario como 1 botella/lote producido
    now = datetime.now()
    fecha = now.strftime("%d/%m/%Y")
    hora = now.strftime("%H:%M:%S")
    
    cursor.execute('''
        INSERT INTO registros (fecha, hora, categoria, producto, cantidad_dictada, botellas_llenas, restante_porcentaje, usuario)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (fecha, hora, receta["categoria"], receta["nombre"], 1.0, 1, "-", f'{user.get("nombre")} (Lab)'))
    
    conn.commit()
    conn.close()
    return {"status": "success", "message": "Lote registrado en inventario"}

# Servir archivos estáticos del frontend en la raíz
app.mount("/", StaticFiles(directory=".", html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=True)
